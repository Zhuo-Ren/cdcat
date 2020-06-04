# -*- coding: utf-8 -*-
from sysdb import SysDb
from openpyxl import load_workbook






class NlpNode(LayerTreeNode):
    def __init__(self):
        # 静态成员
        self.config = None
        self.resource = None
        self.time = None

        # 读取配置文件
        config = {
            type_def: {
                'corpus' : ['absolutePath'], 
                'folder' : ['folderIndex'], 
                'doc' : ['docIndex'],
                'para' : ['paraIndex'],
                'sentence' : ['sentenceIndex'],
                'token' : ['tokenIndex'],
                'char' : ['charIndex']
            }
        } 

        # 初始化层
        for layer_name in config['layers_def'].keys:
            layers[layer_name] = []
            for attr in config['layers_def']['layer_name']:
                self.__setattr__(attr, None)


    def add_type(self, types):
        for type in types:
            if type in LayerTree.type_define
                self.__setattr__()
                self.__setattr__


class NlpNode(LayerTreeNode):
    def __init__():
        source = None
        time - None
        pass


# 有树，有节点。层和配置存在树中
# 节点没有id，使用绝对路径代替id，运行时计算。为：“0-1-0-12-0”这种
# 节点存index：12
# 节点不存左右，只存上下。左右靠id。上下存节点对象。使用列表容器！！！！！！！！！！！
# 层：列表中存节点对象。需要检测层的连续性。
# 节点不设置子类，因为一个节点可能同时是文本、章节、段落，因为一个文本中可以只要一章一段。
# 节点子类改为由类型属性描述，一个节点可以有多个类型。
# 单层类型，如字、词、句、段、章等，对应到layer。
# 如何对应？node.type = {文件夹：{文件名：XXX}}
#          node.type = {
#                       词：{layerIndex：13}，
#                       字：{
#                               layerIndex：98，
#                                text：'看'
#                           } 
#                      }
# 某某类型的特有属性：存放在node.type的对应标签下。

# 标注：
# 库：
# python的树实现

import LayerTree
import LayerTreeNode

LayerTree.config("\config.txt")
LayerTree.root.addNode()


class NlpNode(object):
    '''
    平台中的基本数据结构，一个文件夹是一个node，一个文件是一个node，一句话是一个node，一个句法成分是一个node，一个词是一个node，一个字是一个node。
    node的基本数据成员就是Id和level，对应'nodeTable'中的nodeId和nodeLevel列。
    node根据level的不同，还可能有不同的属性。例如leve=1的node是文件夹，具有name属性。level=2的node是文件，具有content属性。
    '''
    @staticmethod
    def getMaxNodeId():
        return SysDb.getMaxIncrementId('nodeTable')

    @staticmethod
    def getNode(nodeId, nodeName):
        pass

    @staticmethod
    def getNodeById(nodeId):
        # 根据节点ID从数据库中取出该节点的信息
        cmd = "SELECT * FROM nodeTable WHERE nodeId = " + str(nodeId)
        curNode = SysDb.executeCommand(cmd)
        curNode = curNode[0]
        # 将查询到的节点信息赋值给newDict
        nodeTable = SysDb.表结构['nodeTable']
        nodeTabName = nodeTable.keys()
        newDict = dict(zip(nodeTabName, curNode))
        # 根据level实例化一个节点对象
        if newDict['nodeLevel'] == 1:
            newNode = FolderNode(newDict)
        elif newDict['nodeLevel'] == 2:
            newNode = DocNode(newDict)
        elif newDict['nodeLevel'] == 3:
            newNode = NlpNode(newDict)
        elif newDict['nodeLevel'] == 4:
            newNode = ParagraphNode(newDict)
        elif newDict['nodeLevel'] == 5:
            newNode = SentenceNode(newDict)
        elif newDict['nodeLevel'] == 6:
            newNode = ConstituentNode(newDict)
        elif newDict['nodeLevel'] == 7:
            newNode = TokenNode(newDict)
        elif newDict['nodeLevel'] == 8:
            newNode = CharacterNode(newDict)
        else:
            newNode = NlpNode(newDict)
        newNode.nodeId = nodeId
        return newNode

    @staticmethod
    def getNodeIdByIndex(rootId, startIndex, endIndex):
        '''
        :function 指定根节点和查找的开始和结束索引，找到其下面正好涵盖level=词的孩子节点中满足索引范围的孩子的父节点
        :param rootId: Int。以哪个节点作为根节点查找
        :param startIndex: Int。根节点下level=词的孩子节点的起始序号
        :param endIndex: Int。根节点下level=词的孩子节点的终止序号
        :return: 满足条件的节点id
        '''
        rootNode = NlpNode.getNodeById(rootId)
        childIdList = rootNode.getChildNodesId(at=SysDb.levels['词'])  # 词：7；得到root节点在词这一层的所有孩子节点ID
        # endIndex不能超过root所涵盖的最大范围
        if len(childIdList) < endIndex + 1:
            return 'ERROR1: 节点结束索引超出ROOT范围'
        startNode = NlpNode.getNodeById(childIdList[startIndex])  # 对应startIndex的孩子节点
        endNode = NlpNode.getNodeById(childIdList[endIndex])  # 对应endIndex的孩子节点
        parentOfStartNode = startNode.getParentNode()
        while 1:
            # 若startNode的起始内容不等于其父节点的起始内容，说明startNode不是该父节点的第一个孩子，此时应报错
            if parentOfStartNode.contentStart != startNode.contentStart:
                return 'ERROR2：起始index未对齐-起始节点不是所查找节点的第一个符合条件的孩子'
            # 若startNode的结束编号大于endNote的结束编号，说明endNote不是其父节点的最后一个符合条件的孩子，报错
            elif parentOfStartNode.contentEnd > endNode.contentEnd:
                return 'ERROR3：终止index未对齐-小于所查找节点的终止范围'
            # 若startNode的起止编号刚好等于startNode的起始编号和endNote的结束编号，此时符合查找条件，返回该父节点nodeId
            elif parentOfStartNode.contentEnd == endNode.contentEnd:
                return parentOfStartNode.nodeId
            # 若startNode的父节点的结束编号小于endNote的结束编号，说明该父节点未能包括索引范围，应继续往上查找父节点
            elif parentOfStartNode.contentEnd < endNode.contentEnd:
                parentOfStartNode = parentOfStartNode.getParentNode()

    @staticmethod
    def outputToExcel(nodeId, startRow=1, startColumn=1, excelPath='workbook1.xlsx'):
        '''
        :function 把数据库中parentNodeId节点及其以下的outputLevel级节点输出到NlpPlatform/output.xlsx中
        :param nodeId=Int。 需要输出的节点数的根节点
        :param startRow=Int。 从哪一行开始存放数据
        :param startColumn=Int。 从哪一列开始存放数据
        :param excelPath=Str。
        :example 参见 例子：语料+db+excel.rar
        '''

        curNode = NlpNode.getNodeById(nodeId)  # 根据nodeId获取当前节点信息
        childNodeList = curNode.getChildNodes()  # 根据nodeId得到该节点的孩子节点
        wb = load_workbook(excelPath)  # 载入现存excel，每次运行前手动清空
        ws = wb.active
        # 根据节点类型获取单元格里要填入的内容
        if curNode.nodeLevel == SysDb.levels['文件夹']:
            cellContent = curNode.nodeName
        if curNode.nodeLevel == SysDb.levels['文件']:
            cellContent = curNode.nodeName
        if curNode.nodeLevel == SysDb.levels['句']:
            cellContent = '句'
        if curNode.nodeLevel == SysDb.levels['字']:
            cellContent = curNode.getText()
        ws.cell(row=startRow, column=startColumn, value=cellContent)  # 根据行列坐标写excel
        wb.save(excelPath)
        if childNodeList == []:  # 当前节点无孩子，则无需更新
            endRow = startRow
            endColumn = startColumn
        if childNodeList != []:  # 当前节点有孩子，则从其下一行开始存孩子的信息
            childStartColumn = startColumn
            childStartRow = startRow + 1
            # 遍历其子节点
            # 获得子节点的类型
            # 判断childStartRow是否落在子节点类型对应的区域中
            if childNodeList[0].nodeLevel == SysDb.levels['文件夹']:
                if childStartRow < SysDb.levelsZone['文件夹'][0]:  # if 在区域上边
                    childStartRow = SysDb.levelsZone['文件夹'][0]  # childStartRow = 区域的上限
                elif childStartRow > SysDb.levelsZone['文件夹'][1]:  # if 在区域下边
                    ws.insert_rows(childStartRow)  # 插入一行
                    # 更新所有区域上下限
                    delta = childStartRow - SysDb.levelsZone['文件夹'][1]
                    SysDb.levelsZone['文件夹'][1] = childStartRow
                    SysDb.levelsZone['文件'][0] += delta
                    SysDb.levelsZone['文件'][1] += delta
                    SysDb.levelsZone['字'][0] += delta
                    SysDb.levelsZone['字'][1] += delta
            elif childNodeList[0].nodeLevel == SysDb.levels['文件']:
                if childStartRow < SysDb.levelsZone['文件'][0]:  # if 在区域上边
                    childStartRow = SysDb.levelsZone['文件'][0]  # childStartRow = 区域的上限
                elif childStartRow > SysDb.levelsZone['文件'][1]:  # if 在区域下边
                    ws.insert_rows(childStartRow)  # 插入一行
                    # 更新
                    delta = childStartRow - SysDb.levelsZone['文件'][1]
                    SysDb.levelsZone['文件'][0] += delta
                    SysDb.levelsZone['文件'][1] += delta
                    SysDb.levelsZone['字'][0] += delta
                    SysDb.levelsZone['字'][1] += delta
            elif childNodeList[0].nodeLevel == SysDb.levels['字']:
                if childStartRow < SysDb.levelsZone['字'][0]:  # if 在区域上边
                    childStartRow = SysDb.levelsZone['字'][0]  # childStartRow = 区域的上限
                elif childStartRow > SysDb.levelsZone['字'][1]:  # if 在区域下边
                    ws.insert_rows(childStartRow)  # 插入一行
                    # 更新
                    delta = childStartRow - SysDb.levelsZone['字'][1]
                    SysDb.levelsZone['字'][0] += delta
                    SysDb.levelsZone['字'][1] += delta
                    # SysDb.levelsZone['字'][0]
            # 第一个孩子节点和后续其兄弟节点分开讨论
            childEndRow, childEndColumn = NlpNode.outputToExcel(childNodeList[0].nodeId, childStartRow, childStartColumn, excelPath=excelPath)
            for childNode in childNodeList[1:]:
                childStartRow = childEndRow
                childStartColumn = childEndColumn + 1
                # 获得兄节点的类型
                if childNode.nodeLevel == SysDb.levels['文件夹']:
                    if childStartRow < SysDb.levelsZone['文件夹'][0]:  # if 在区域上边
                        childStartRow = SysDb.levelsZone['文件夹'][0]  # childStartRow = 区域的上限
                    elif childStartRow > SysDb.levelsZone['文件夹'][1]:  # if 在区域下边
                        ws.insert_rows(childStartRow)  # 插入一行
                        # 更新所有区域上下限
                        delta = childStartRow - SysDb.levelsZone['文件夹'][1]
                        SysDb.levelsZone['文件夹'][1] = childStartRow
                        SysDb.levelsZone['文件'][0] += delta
                        SysDb.levelsZone['文件'][1] += delta
                        SysDb.levelsZone['字'][0] += delta
                        SysDb.levelsZone['字'][1] += delta
                elif childNode.nodeLevel == SysDb.levels['文件']:
                    if childStartRow < SysDb.levelsZone['文件'][0]:  # if 在区域上边
                        childStartRow = SysDb.levelsZone['文件'][0]  # childStartRow = 区域的上限
                    elif childStartRow > SysDb.levelsZone['文件'][1]:  # if 在区域下边
                        ws.insert_rows(childStartRow)  # 插入一行
                        # 更新
                        delta = childStartRow - SysDb.levelsZone['文件'][1]
                        SysDb.levelsZone['文件'][0] += delta
                        SysDb.levelsZone['文件'][1] += delta
                        SysDb.levelsZone['字'][0] += delta
                        SysDb.levelsZone['字'][1] += delta
                elif childNode.nodeLevel == SysDb.levels['字']:
                    if childStartRow < SysDb.levelsZone['字'][0]:  # if 在区域上边
                        childStartRow = SysDb.levelsZone['字'][0]  # childStartRow = 区域的上限
                    elif childStartRow > SysDb.levelsZone['字'][1]:  # if 在区域下边
                        ws.insert_rows(childStartRow)  # 插入一行
                        # 更新
                        delta = childStartRow - SysDb.levelsZone['字'][1]
                        SysDb.levelsZone['字'][0] += delta
                        SysDb.levelsZone['字'][1] += delta
                childEndRow, childEndColumn = NlpNode.outputToExcel(childNode.nodeId, childStartRow, childStartColumn, excelPath=excelPath)
            endRow = startRow  # 当前节点的end坐标一定和其start坐标同行
            endColumn = childEndColumn
        # 合并单元格
        wb = load_workbook(excelPath)
        ws = wb.active
        ws.merge_cells(start_row=startRow, start_column=startColumn, end_row=endRow, end_column=endColumn)
        wb.save(excelPath)
        return endRow, endColumn

    def __init__(self, propDict={}):
        # 初始化动态数据成员
        self.nodeId = SysDb.表结构['nodeTable']['nodeId']['初始值']
        '''
        注：nodeId为初始值，则意味着此节点对象从未被存入nodeTable；nodeId为数字，则意味着存入过nodeTable
        '''
        self.nodeLevel = SysDb.表结构['nodeTable']['nodeLevel']['初始值']
        self.leftNodeId = SysDb.表结构['nodeTable']['leftNodeId']['初始值']
        self.rightNodeId = SysDb.表结构['nodeTable']['rightNodeId']['初始值']
        self.parentNodeId = SysDb.表结构['nodeTable']['parentNodeId']['初始值']
        
        # 设置动态数据成员
        self.setProps(propDict)

    def setProps(self, propDict):
        if 'nodeLevel' in list(propDict.keys()):
            self.nodeLevel = propDict['nodeLevel']
        if 'leftNodeId' in list(propDict.keys()):
            self.leftNodeId = propDict['leftNodeId'] 
        if 'rightNodeId' in list(propDict.keys()):
            self.rightNodeId = propDict['rightNodeId'] 
        if 'parentNodeId' in list(propDict.keys()):
            self.parentNodeId = propDict['parentNodeId'] 

    def getProps(self, propsType='all'):
        nodeTabledict = {
            'nodeId': self.nodeId,
            'nodeLevel': self.nodeLevel,
            'leftNodeId': self.leftNodeId,
            'rightNodeId': self.rightNodeId,
            'parentNodeId': self.parentNodeId
        }
        return nodeTabledict

    def saveNode(self):
        """
        :function 把此节点信息同步到数据库中“nodeTable”。
            如果是新节点就在nodeTable中创建一行记录，如果是已有节点就修改nodeTable中对应记录。
        :propDict 节点属性以字典形式传递，没有可以不传
        :returns 新存入的这个节点的nodeId
        """
        if self.nodeId == SysDb.表结构['nodeTable']['nodeId']['初始值']:
            self.saveNodeCreate()
        else:
            self.saveNodeChange()

    def saveNodeCreate(self):
        """
        """
        # 生成一个nodeTable全属性初始值dict作为基础
        propsDict = {k: SysDb.表结构['nodeTable'][k]['初始值'] for k in list(SysDb.表结构['nodeTable'].keys())}
        # 加入本节点的信息
        propsDict.update(self.getProps(propsType='nodeTable'))
        # 新节点写入数据库
        SysDb.insertRow('nodeTable', propsDict)
        # 入库完成后，更新nodeId
        self.nodeId = SysDb.getMaxIncrementId('nodeTable')

    def saveNodeChange(self):
        """
        """
        # 把节点的更新入库
        SysDb.updateRow('nodeTable', {'nodeId': self.nodeId}, self.getProps(propsType='nodeTable'))

    def getChildNodesId(self, at=None, higher=None, lower=None):
        # 获取直接孩子
        if self.nodeId is None:
            return []
        else:
            cmd = "SELECT nodeId FROM nodeTable WHERE parentNodeId = " + str(self.nodeId)
            directChildNodesIdList = SysDb.executeCommand(cmd)  # 此时返回值类似于：[(2,), (3,)]
            directChildNodesIdList = [x[0] for x in directChildNodesIdList]  # 此时返回值类似于：[2,3]
        # 处理模式1：无参
        if at is None and higher is None and lower is None:
            return directChildNodesIdList
        # 处理模式2：at
        elif at is not None:
            # 输入检测
            if at < self.nodeLevel:
                return "非法的输入参数：at<nodeLevel"
            childNodesIdList = []
            for directChildId in directChildNodesIdList:
                # 根据直接孩子ID，获取直接孩子节点
                directChildNode = NlpNode.getNodeById(directChildId)
                # 未达域->深入：如果直接孩子未到达,尝试更深层次的遍历
                if directChildNode.nodeLevel < at:
                    # 尝试更深层次的遍历
                    deeper = directChildNode.getChildNodesId(at=at)
                    # 如果尝试不成功，就说明这个直接孩子非法。
                    #   只要有任何一个直接孩子是非法的，那么整个直接孩子列表都是非法的，返回[]
                    if deeper == []:
                        return []
                    # 如果尝试成功，输出尝试结果。
                    else:
                        childNodesIdList.extend(deeper)
                # 合法域->输出：at模式是懒惰的，只要直接孩子合法，直接输出，不继续遍历
                elif directChildNode.nodeLevel == at:
                    childNodesIdList.append(directChildId)
                # 非法域：只要有任何一个直接孩子是非法的，那么整个直接孩子列表都是非法的，返回[]
                elif at < directChildNode.nodeLevel:
                    return []
        # 处理模式3：上界
        elif at is None and higher is not None:
            # 输入检测
            if higher < self.nodeLevel:
                return "非法的输入参数：higher<nodeLevel"
            childNodesIdList = []
            for directChildId in directChildNodesIdList:
                # 根据直接孩子ID，获取直接孩子节点
                directChildNode = NlpNode.getNodeById(directChildId)
                # 未达域->深入：如果直接孩子未到达,尝试更深层次的遍历
                if directChildNode.nodeLevel < higher:
                    # 尝试更深层次的遍历
                    deeper = directChildNode.getChildNodesId(higher=higher)
                    # 如果尝试不成功，就说明这个直接孩子非法。
                    #   只要有任何一个直接孩子是非法的，那么整个直接孩子列表都是非法的，返回[]
                    if deeper == []:
                        return []
                    # 如果尝试成功，输出尝试结果。
                    else:
                        childNodesIdList.extend(deeper)
                # 合法域->输出：higher模式是懒惰的，只要直接孩子合法，直接输出，不继续遍历
                elif higher <= directChildNode.nodeLevel:
                    childNodesIdList.append(directChildId)
        # 处理模式4:下界
        elif at is None and higher is None and lower is not None:
            # 输入检测
            if lower < self.nodeLevel:
                return "非法的输入参数：lower<nodeLevel"
            childNodesIdList = []
            for directChildId in directChildNodesIdList:
                # 根据直接孩子ID，获取直接孩子节点
                directChildNode = NlpNode.getNodeById(directChildId)
                # 合法域->深入：lower模式是勤奋的，总是尝试继续遍历
                if directChildNode.nodeLevel <= lower:
                    # 尝试更深层次的遍历
                    deeper = directChildNode.getChildNodesId(lower=lower)
                    # 如果尝试不成功，就抛弃尝试结果，输出直接孩子
                    if deeper == []:
                        childNodesIdList.append(directChildId)
                    # 如果尝试成功，输出尝试结果。
                    else:
                        childNodesIdList.extend(deeper)
                # 非法域：只要有任何一个直接孩子是非法的，那么整个直接孩子列表都是非法的，返回[]
                elif lower < directChildNode.nodeLevel:
                    return []
        return childNodesIdList

    def getChildNodes(self, at=None, higher=None, lower=None):
        if self.nodeId is None:
            return []
        else:
            ChildNodesIdList = self.getChildNodesId(at=at, higher=higher, lower=lower)
            childNodesList = [NlpNode.getNodeById(x) for x in ChildNodesIdList]
            return childNodesList

    def getRightNodeId(self):
        return self.rightNodeId

    def getRightNode(self):
        if self.nodeId == None:
            return None
        else:
            return NlpNode.getNodeById(self.rightNodeId)

    def getNthRightNodeId(self, n):
        # if self.nodeId == None:
        #     return None
        # else: 
        #     curNodeId = self.nodeId
        #     for i in range(1,n):
        #         cmd = "SELECT rightNodeId FROM nodeTable WHERE NodeId = " + str(curNodeId)
        #         curNodeId = SysDb.executeCommand(cmd)
        #     return curNodeId
        pass

    def getNthChildNode(self, n):
        # NthChildNodeId = self.getNthChildNodeId(n)
        # return NlpNode.getNodeById(NthChildNodeId)
        pass

    def getNthChildNodeId(self):
        # childNode = self.getChildNodesId
        # return childNode[n]
        pass

    def getParentNode(self):
        if self.parentNodeId != None :
            return NlpNode.getNodeById(self.parentNodeId)
        else:
            return None

    def getParentNodeId(self, at=None, higher=None, lower=None):
        if self.nodeId is None:
            return None
        else:
            directParentNodeId = self.parentNodeId
        # 处理模式1：无参
        if at is None and higher is None and lower is None:
            return directParentNodeId
        # 处理模式2：at
        elif at is not None:  # 优先级最高。给定后返回子节点给定层级的父节点ID
            parentNode = NlpNode.getNodeById(self.parentNodeId)
            while parentNode.nodeLevel != at:
                curParentId = parentNode.nodeId
                parentNode = NlpNode.getNodeById(parentNode.parentNodeId)
                if parentNode.nodeLevel < at:
                    return None
                # elif parentNode.nodeId == '根':
                #     return curParentId
            return parentNode.nodeId
        # 处理模式3：下界
        elif at is None and lower is not None:  # 优先级其次。给定后只能在包括lower层以及其以上的层级给出离给定节点最近的父节点ID
            if lower < 1:
                return "非法的输入参数：lower<1"
            parentNode = NlpNode.getNodeById(self.parentNodeId)
            while parentNode.nodeLevel > lower:
                getter = parentNode.nodeLevel
                parentNode = NlpNode.getNodeById(parentNode.parentNodeId)
            return parentNode.nodeId
        # 处理模式4:上界
        elif at is None and lower is None and higher is not None:  # 优先级最低。给定后只能在包括lower层以及其以下的层级给出离给定节点最近的父节点ID
            if higher > self.nodeLevel:
                # print("非法的输入参数：higher>nodeLevel")
                return None
            parentNode = NlpNode.getNodeById(self.parentNodeId)
            if parentNode.nodeLevel < higher:
                return None
            while parentNode.nodeLevel > higher:
                curID = parentNode.parentNodeId
                parentNode = NlpNode.getNodeById(parentNode.parentNodeId)
            while parentNode.nodeLevel == higher and parentNode.parentNodeId != '根':
                curParentId = parentNode.nodeId
                parentNode = NlpNode.getNodeById(parentNode.parentNodeId)
                if parentNode.nodeLevel < higher:
                    return curParentId
            return parentNode.nodeId

    def getText(self):
        # 若当前节点为文件夹节点，则需要往下得到文件夹内所有文件的nodeContent
        if self.nodeLevel == SysDb.levels['文件夹']:
            docNodeList = self.getChildNodes(at=SysDb.levels['文件'])
            text = ''
            for docNode in docNodeList:
                text = text + docNode.getText()
            return text
        # 若当前节点是文件节点：直接取nodeContent
        elif self.nodeLevel == SysDb.levels['文件']:
            return self.nodeContent
        # 若当前节点是更下级节点。取文件的下级内容（如章、段、句、字等）：一直往上取到包含其的文件节点中的nodeContent
        else:
            curNode = self
            # parentNode = NlpNode.getNodeById(curNode.parentNodeId)
            while curNode.nodeLevel != SysDb.levels['文件']:
                curNode = NlpNode.getNodeById(curNode.parentNodeId)
            curNodeContent = curNode.nodeContent
            charStart = int(self.contentStart)
            charEnd = int(self.contentEnd) + 1
            charText = curNodeContent[charStart:charEnd]
            return charText


class FolderNode(NlpNode):
    def __init__(self, propDict={}):
        # 初始化动态数据成员
        self.nodeId = SysDb.表结构['nodeTable']['nodeId']['初始值']
        self.nodeLevel = SysDb.levels['文件夹']
        self.leftNodeId = SysDb.表结构['nodeTable']['leftNodeId']['初始值']
        self.rightNodeId = SysDb.表结构['nodeTable']['rightNodeId']['初始值']
        self.parentNodeId = SysDb.表结构['nodeTable']['parentNodeId']['初始值']
        self.nodeName = SysDb.表结构['nodeTable']['nodeName']['初始值']

        # 设置动态数据成员
        self.setProps(propDict)

    def setProps(self, propDict):
        if 'leftNodeId' in list(propDict.keys()):
            self.leftNodeId = propDict['leftNodeId'] 
        if 'rightNodeId' in list(propDict.keys()):
            self.rightNodeId = propDict['rightNodeId'] 
        if 'parentNodeId' in list(propDict.keys()):
            self.parentNodeId = propDict['parentNodeId']
        if 'nodeName' in list(propDict.keys()):
            self.nodeName = propDict['nodeName'] 

    def getProps(self, propsType='all'):
        nodeTabledict = {
            'nodeId': self.nodeId,
            'nodeLevel': self.nodeLevel,
            'leftNodeId': self.leftNodeId,
            'rightNodeId': self.rightNodeId,
            'parentNodeId': self.parentNodeId,
            'nodeName': self.nodeName
        }
        return nodeTabledict


class DocNode(NlpNode):
    def __init__(self, propDict={}):
        # 初始化动态数据成员
        self.nodeId = SysDb.表结构['nodeTable']['nodeId']['初始值']
        self.nodeLevel = SysDb.levels['文件']
        self.leftNodeId = SysDb.表结构['nodeTable']['leftNodeId']['初始值']
        self.rightNodeId = SysDb.表结构['nodeTable']['rightNodeId']['初始值']
        self.parentNodeId = SysDb.表结构['nodeTable']['parentNodeId']['初始值']
        self.nodeName = SysDb.表结构['nodeTable']['nodeName']['初始值']
        self.nodeContent = SysDb.表结构['nodeTable']['nodeContent']['初始值']
        self.contentStart = 0
        self.contentEnd = SysDb.表结构['nodeTable']['contentEnd']['初始值']

        # 设置动态数据成员
        self.setProps(propDict)

    def setProps(self, propDict):
        if 'leftNodeId' in list(propDict.keys()):
            self.leftNodeId = propDict['leftNodeId'] 
        if 'rightNodeId' in list(propDict.keys()):
            self.rightNodeId = propDict['rightNodeId'] 
        if 'parentNodeId' in list(propDict.keys()):
            self.parentNodeId = propDict['parentNodeId']
        if 'nodeName' in list(propDict.keys()):
            self.nodeName = propDict['nodeName']
        if 'nodeContent' in list(propDict.keys()):
            self.nodeContent = propDict['nodeContent']
            self.contentEnd = len(self.nodeContent) - 1

    def getProps(self, propsType='all'):
        nodeTabledict = {
            'nodeId': self.nodeId,
            'nodeLevel': self.nodeLevel,
            'leftNodeId': self.leftNodeId,
            'rightNodeId': self.rightNodeId,
            'parentNodeId': self.parentNodeId,
            'nodeName': self.nodeName,
            'nodeContent': self.nodeContent,
            'contentStart': self.contentStart,
            'contentEnd': self.contentEnd
        }
        return nodeTabledict


class ParagraphNode(NlpNode):
    def __init__(self, propDict={}):
        # 初始化动态数据成员
        self.nodeId = SysDb.表结构['nodeTable']['nodeId']['初始值']
        self.nodeLevel = SysDb.levels['段']
        self.leftNodeId = SysDb.表结构['nodeTable']['leftNodeId']['初始值']
        self.rightNodeId = SysDb.表结构['nodeTable']['rightNodeId']['初始值']
        self.parentNodeId = SysDb.表结构['nodeTable']['parentNodeId']['初始值']
        self.contentStart = SysDb.表结构['nodeTable']['contentStart']['初始值']
        self.contentEnd = SysDb.表结构['nodeTable']['contentEnd']['初始值']

        # 设置动态数据成员
        self.setProps(propDict)

    def setProps(self, propDict):
        if 'leftNodeId' in list(propDict.keys()):
            self.leftNodeId = propDict['leftNodeId']
        if 'rightNodeId' in list(propDict.keys()):
            self.rightNodeId = propDict['rightNodeId']
        if 'parentNodeId' in list(propDict.keys()):
            self.parentNodeId = propDict['parentNodeId']
        if 'contentStart' in list(propDict.keys()):
            self.contentStart = propDict['contentStart']
        if 'contentEnd' in list(propDict.keys()):
            self.contentEnd = propDict['contentEnd']

    def getProps(self, propsType='all'):
        nodeTabledict = {
            'nodeId': self.nodeId,
            'nodeLevel': self.nodeLevel,
            'leftNodeId': self.leftNodeId,
            'rightNodeId': self.rightNodeId,
            'parentNodeId': self.parentNodeId,
            'contentStart': self.contentStart,
            'contentEnd': self.contentEnd
        }
        return nodeTabledict


class SentenceNode(NlpNode):
    def __init__(self, propDict={}):
        # 初始化动态数据成员
        self.nodeId = SysDb.表结构['nodeTable']['nodeId']['初始值']
        self.nodeLevel = SysDb.levels['句']
        self.leftNodeId = SysDb.表结构['nodeTable']['leftNodeId']['初始值']
        self.rightNodeId = SysDb.表结构['nodeTable']['rightNodeId']['初始值']
        self.parentNodeId = SysDb.表结构['nodeTable']['parentNodeId']['初始值']
        self.contentStart = SysDb.表结构['nodeTable']['contentStart']['初始值']
        self.contentEnd = SysDb.表结构['nodeTable']['contentEnd']['初始值']

        # 设置动态数据成员
        self.setProps(propDict)

    def setProps(self, propDict):
        if 'leftNodeId' in list(propDict.keys()):
            self.leftNodeId = propDict['leftNodeId']
        if 'rightNodeId' in list(propDict.keys()):
            self.rightNodeId = propDict['rightNodeId']
        if 'parentNodeId' in list(propDict.keys()):
            self.parentNodeId = propDict['parentNodeId']
        if 'contentStart' in list(propDict.keys()):
            self.contentStart = propDict['contentStart']
        if 'contentEnd' in list(propDict.keys()):
            self.contentEnd = propDict['contentEnd']

    def getProps(self, propsType='all'):
        nodeTabledict = {
            'nodeId': self.nodeId,
            'nodeLevel': self.nodeLevel,
            'leftNodeId': self.leftNodeId,
            'rightNodeId': self.rightNodeId,
            'parentNodeId': self.parentNodeId,
            'contentStart': self.contentStart,
            'contentEnd': self.contentEnd
        }
        return nodeTabledict

    # def getText(self):
    #     parentNode = NlpNode.getNodeById(self.parentNodeId)
    #     if parentNode.nodeLevel != 2:
    #         parentNode = NlpNode.getNodeById(parentNode.parentNodeId)
    #     parentContent = parentNode.nodeContent
    #     charStart = int(self.contentStart)
    #     charEnd = int(self.contentEnd) + 1
    #     charText = parentContent[charStart:charEnd]
    #     return charText


class ConstituentNode(NlpNode):
    def __init__(self, propDict={}):
        # 初始化动态数据成员
        self.nodeId = SysDb.表结构['nodeTable']['nodeId']['初始值']
        self.nodeLevel = SysDb.levels['成分']
        self.leftNodeId = SysDb.表结构['nodeTable']['leftNodeId']['初始值']
        self.rightNodeId = SysDb.表结构['nodeTable']['rightNodeId']['初始值']
        self.parentNodeId = SysDb.表结构['nodeTable']['parentNodeId']['初始值']
        self.contentStart = SysDb.表结构['nodeTable']['contentStart']['初始值']
        self.contentEnd = SysDb.表结构['nodeTable']['contentEnd']['初始值']

        # 设置动态数据成员
        self.setProps(propDict)

    def setProps(self, propDict):
        if 'leftNodeId' in list(propDict.keys()):
            self.leftNodeId = propDict['leftNodeId']
        if 'rightNodeId' in list(propDict.keys()):
            self.rightNodeId = propDict['rightNodeId']
        if 'parentNodeId' in list(propDict.keys()):
            self.parentNodeId = propDict['parentNodeId']
        if 'contentStart' in list(propDict.keys()):
            self.contentStart = propDict['contentStart']
        if 'contentEnd' in list(propDict.keys()):
            self.contentEnd = propDict['contentEnd']

    def getProps(self, propsType='all'):
        nodeTabledict = {
            'nodeId': self.nodeId,
            'nodeLevel': self.nodeLevel,
            'leftNodeId': self.leftNodeId,
            'rightNodeId': self.rightNodeId,
            'parentNodeId': self.parentNodeId,
            'contentStart': self.contentStart,
            'contentEnd': self.contentEnd
        }
        return nodeTabledict


class TokenNode(NlpNode):
    def __init__(self, propDict={}):
        # 初始化动态数据成员
        self.nodeId = SysDb.表结构['nodeTable']['nodeId']['初始值']
        self.nodeLevel = SysDb.levels['词']
        self.leftNodeId = SysDb.表结构['nodeTable']['leftNodeId']['初始值']
        self.rightNodeId = SysDb.表结构['nodeTable']['rightNodeId']['初始值']
        self.parentNodeId = SysDb.表结构['nodeTable']['parentNodeId']['初始值']
        self.contentStart = SysDb.表结构['nodeTable']['contentStart']['初始值']
        self.contentEnd = SysDb.表结构['nodeTable']['contentEnd']['初始值']

        # 设置动态数据成员
        self.setProps(propDict)

    def setProps(self, propDict):
        if 'leftNodeId' in list(propDict.keys()):
            self.leftNodeId = propDict['leftNodeId']
        if 'rightNodeId' in list(propDict.keys()):
            self.rightNodeId = propDict['rightNodeId']
        if 'parentNodeId' in list(propDict.keys()):
            self.parentNodeId = propDict['parentNodeId']
        if 'contentStart' in list(propDict.keys()):
            self.contentStart = propDict['contentStart']
        if 'contentEnd' in list(propDict.keys()):
            self.contentEnd = propDict['contentEnd']

    def getProps(self, propsType='all'):
        nodeTabledict = {
            'nodeId': self.nodeId,
            'nodeLevel': self.nodeLevel,
            'leftNodeId': self.leftNodeId,
            'rightNodeId': self.rightNodeId,
            'parentNodeId': self.parentNodeId,
            'contentStart': self.contentStart,
            'contentEnd': self.contentEnd
        }
        return nodeTabledict


class CharacterNode(NlpNode):
    def __init__(self, propDict={}):
        # 初始化动态数据成员
        self.nodeId = SysDb.表结构['nodeTable']['nodeId']['初始值']
        self.nodeLevel = SysDb.levels['字']
        self.leftNodeId = SysDb.表结构['nodeTable']['leftNodeId']['初始值']
        self.rightNodeId = SysDb.表结构['nodeTable']['rightNodeId']['初始值']
        self.parentNodeId = SysDb.表结构['nodeTable']['parentNodeId']['初始值']
        self.contentStart = SysDb.表结构['nodeTable']['contentStart']['初始值']
        self.contentEnd = SysDb.表结构['nodeTable']['contentEnd']['初始值']

        # 设置动态数据成员
        self.setProps(propDict)

    def setProps(self, propDict):
        if 'leftNodeId' in list(propDict.keys()):
            self.leftNodeId = propDict['leftNodeId'] 
        if 'rightNodeId' in list(propDict.keys()):
            self.rightNodeId = propDict['rightNodeId'] 
        if 'parentNodeId' in list(propDict.keys()):
            self.parentNodeId = propDict['parentNodeId']
        if 'contentStart' in list(propDict.keys()):
            self.contentStart = propDict['contentStart']
        if 'contentEnd' in list(propDict.keys()):
            self.contentEnd = propDict['contentEnd']

    def getProps(self, propsType='all'):
        nodeTabledict = {
            'nodeId': self.nodeId,
            'nodeLevel': self.nodeLevel,
            'leftNodeId': self.leftNodeId,
            'rightNodeId': self.rightNodeId,
            'parentNodeId': self.parentNodeId,
            'contentStart': self.contentStart,
            'contentEnd': self.contentEnd
        }
        return nodeTabledict